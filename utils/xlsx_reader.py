# --*-- coding: utf-8 --*--
import sys
import os

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__),".."))
sys.path.append(BASE_DIR)

import openpyxl
from db_access import create_company, create_website

wb = openpyxl.load_workbook('../docs/swiftwire.xlsx')

def sheet01():
    sheet = wb.get_sheet_by_name("原有公司")
    for item in sheet.iter_rows(min_row=2, max_col=12, max_row=400):
        if not item[1].value:
            continue

        com_info = {
            'name_cn': item[1].value,
            'name_en': item[2].value,
            'industry': item[3].value
        }

        res = create_company(**com_info)
        company = res["company"]
        if company:
            for i in range(4, 11):
                if item[i].value:
                    site_info = {
                        'url': item[i].value,
                        'company_id': company.id,
                    }
                    create_website(**site_info)


def sheet02():
    sheet = wb.get_sheet_by_name("没有重复的公司")
    for item in sheet.iter_rows(min_row=2, max_col=12, max_row=400):
        if not item[1].value:
            continue

        com_info = {
            'name_cn': item[1].value,
            'name_en': item[2].value,
            'industry': item[3].value
        }

        res = create_company(**com_info)
        company = res["company"]
        if company:
            for i in range(4, 11):
                if item[i].value:
                    site_info = {
                        'url': item[i].value,
                        'company_id': company.id,
                    }
                    create_website(**site_info)


def sheet03():
    sheet = wb.get_sheet_by_name("新创公司")
    for item in sheet.iter_rows(min_row=3, max_col=12, max_row=400):
        if not item[1].value:
            continue

        com_info = {
            'name_cn': item[1].value,
            'name_en': '',
            'industry': ''
        }

        res = create_company(**com_info)
        company = res["company"]
        if company:
            if item[2].value:
                site_info = {
                    'url': item[2].value,
                    'company_id': company.id,
                }
                create_website(**site_info)



if __name__ == '__main__':
    sheet01()
    sheet02()
    sheet03()

